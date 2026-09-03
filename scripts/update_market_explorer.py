import argparse
import json
from pathlib import Path

import openpyxl


def split_materials(value: object) -> list[str]:
    text = str(value or '').strip()
    if not text:
        return []
    parts: list[str] = []
    start = 0
    depth = 0
    pairs = {'(': ')', '[': ']', '{': '}'}
    closing = set(pairs.values())
    for index, char in enumerate(text):
        if char in pairs:
            depth += 1
        elif char in closing:
            depth = max(0, depth - 1)
        elif char == ',' and depth == 0:
            part = text[start:index].strip()
            if part:
                parts.append(part)
            start = index + 1
    final = text[start:].strip()
    if final:
        parts.append(final)
    return parts


def number(value: object) -> int | float:
    numeric = float(value or 0)
    return int(numeric) if numeric.is_integer() else numeric


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Convert the MFDS production-sales workbook into the browser dataset."
    )
    parser.add_argument("source", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook = openpyxl.load_workbook(args.source, data_only=True, read_only=True)
    trend_sheet = workbook['인정형태별_추이']
    trend_rows = list(trend_sheet.iter_rows(values_only=True))
    years = [int(value) for value in trend_rows[0][1:] if value is not None]
    data = {
        'trend': {
            'years': years,
            '전체': [number(value) for value in trend_rows[1][1:1 + len(years)]],
            '고시형': [number(value) for value in trend_rows[3][1:1 + len(years)]],
            '개별인정형': [number(value) for value in trend_rows[5][1:1 + len(years)]],
        },
        'func': {str(year): {category: [] for category in ('전체', '고시형', '개별인정형')} for year in (2023, 2024, 2025)},
        'ing': {str(year): {category: [] for category in ('고시형', '개별인정형')} for year in (2023, 2024, 2025)},
    }

    function_sheets = {
        '전체': '기능성별_전체',
        '고시형': '기능성별_고시형',
        '개별인정형': '기능성별_개별인정형',
    }
    for category, sheet_name in function_sheets.items():
        rows = workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row[0] or not row[1] or str(row[1]).strip() in {'계', '합계'}:
                continue
            year = str(int(row[0]))
            entry = {
                'name': str(row[1]).strip(),
                'val': number(row[2]),
                'pct': round(float(row[3] or 0) * 100, 3),
            }
            if category != '전체':
                entry['mat'] = split_materials(row[4])
            data['func'][year][category].append(entry)

    ingredient_sheets = {
        '고시형': '원료별_고시형',
        '개별인정형': '원료별_개별인정형',
    }
    for category, sheet_name in ingredient_sheets.items():
        rows = workbook[sheet_name].iter_rows(min_row=2, values_only=True)
        for row in rows:
            if not row[0] or row[1] is None or not row[2]:
                continue
            year = str(int(row[0]))
            data['ing'][year][category].append({
                'rank': int(row[1]),
                'name': str(row[2]).strip(),
                'tot': number(row[3]),
                'dom': number(row[4]),
                'exp': number(row[5]),
            })

    payload = json.dumps(data, ensure_ascii=False, separators=(",", ":"))
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        "window.MARKET_EXPLORER_DATA=" + payload + ";\n",
        encoding="utf-8",
    )


if __name__ == "__main__":
    main()
